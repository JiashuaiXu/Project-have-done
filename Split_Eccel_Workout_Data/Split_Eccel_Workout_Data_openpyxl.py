import shutil
from openpyxl import load_workbook

# 定义一个函数，复制文件并保留对应周的工作表
def copy_and_filter_workbook(week_num, input_filename, output_filename):
    # 复制文件
    shutil.copy(input_filename, output_filename)
    
    # 加载新复制的文件
    workbook = load_workbook(output_filename)
    
    # 获取所有工作表名称
    worksheet_names = workbook.sheetnames
    
    # 找出需要保留的工作表
    week_worksheets = [ws for ws in worksheet_names if f'Week_{week_num}' in ws]
    
    # 删除不需要的工作表
    for ws_name in worksheet_names:
        if ws_name not in week_worksheets:
            workbook.remove(workbook[ws_name])
    
    # 保存修改后的工作簿
    workbook.save(output_filename)

# 输入文件名
input_filename = '训练记录_week31-41.xlsx'

# 拆分并保存所有周（从31到41）
for week in range(31, 42):
    output_filename = f'训练记录_week{week}.xlsx'
    copy_and_filter_workbook(week, input_filename, output_filename)
